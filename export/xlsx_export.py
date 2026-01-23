"""
XLSX Export functionality for test plans.

This module provides functions to export test plan data to XLSX format,
making it easy to work with in tools like Excel.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Union

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # type: ignore

from database.models import TestTaskModel
from tasks.test_task_model import TestTask


def _extract_task_data(task: Union[TestTask, TestTaskModel, Dict[str, Any]]) -> Dict[str, Any]:
    """
    Extract task data from various input types for XLSX export.
    
    Args:
        task: Task object (TestTask, TestTaskModel) or dictionary
        
    Returns:
        Dictionary with task data
    """
    if isinstance(task, dict):
        return task
    
    # Handle TestTask dataclass
    if isinstance(task, TestTask):
        return {
            'id': task.id,
            'description': task.description,
            'owner': task.owner or '',
            'coverage_status': getattr(task.coverage_status, 'value', str(task.coverage_status)),
            'status': getattr(task.status, 'value', str(task.status)),
            'test_type': getattr(task.test_type, 'value', str(task.test_type)),
            'dependencies': ', '.join(task.dependencies) if task.dependencies else '',
        }
    
    # Handle TestTaskModel (SQLAlchemy model)
    if isinstance(task, TestTaskModel):
        return {
            'id': task.id,
            'description': task.description,
            'owner': task.owner or '',
            'coverage_status': getattr(task.coverage_status, 'value', str(task.coverage_status)),
            'status': getattr(task.status, 'value', str(task.status)),
            'test_type': getattr(task.test_type, 'value', str(task.test_type)),
            'dependencies': ', '.join(task.dependencies) if task.dependencies else '',
        }
    
    raise TypeError(f"Unsupported task type: {type(task)}")


def _format_worksheet(ws: Any, num_rows: int, num_cols: int) -> None:
    """
    Format the worksheet for better readability.
    
    Args:
        ws: openpyxl worksheet object
        num_rows: Number of data rows (excluding header)
        num_cols: Number of columns
    """
    if not OPENPYXL_AVAILABLE:
        return
    
    # Format header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Auto-adjust column widths
    for col in range(1, num_cols + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        # Check header
        header_cell = ws.cell(row=1, column=col)
        if header_cell.value:
            max_length = len(str(header_cell.value))
        
        # Check data rows
        for row in range(2, num_rows + 2):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set column width (add some padding)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Freeze header row
    ws.freeze_panes = "A2"


def export_to_xlsx(
    test_plan: Iterable[Union[TestTask, TestTaskModel, Dict[str, Any]]],
    output_path: Union[str, Path],
    include_headers: bool = True,
    columns: Optional[List[str]] = None,
    format_sheet: bool = True,
) -> Path:
    """
    Export test plan data to XLSX format.
    
    Args:
        test_plan: Iterable of task objects (TestTask, TestTaskModel) or dictionaries
        output_path: Path where the XLSX file should be written
        include_headers: Whether to include column headers (default: True)
        columns: Optional list of column names to include. If None, uses default columns.
                 Default columns: ['Task ID', 'Description', 'Owner', 'Coverage Status',
                                  'Status', 'Test Type', 'Dependencies']
        format_sheet: Whether to apply formatting (colors, widths, etc.) (default: True)
    
    Returns:
        Path object pointing to the created XLSX file
        
    Raises:
        ImportError: If openpyxl is not installed
        
    Example:
        >>> from tasks.test_plan_generator import TestPlanGenerator
        >>> from database.data_access_layer import TestTaskDAL
        >>> 
        >>> generator = TestPlanGenerator()
        >>> dal = TestTaskDAL(session)
        >>> plan = generator.generate_plan_from_db(dal)
        >>> 
        >>> export_to_xlsx(plan, 'test_plan.xlsx')
        Path('test_plan.xlsx')
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for XLSX export. Install it with: pip install openpyxl"
        )
    
    output_path = Path(output_path)
    
    # Default columns if not specified
    if columns is None:
        columns = [
            'Task ID',
            'Description',
            'Owner',
            'Coverage Status',
            'Status',
            'Test Type',
            'Dependencies',
        ]
    
    # Map column names to task data keys
    column_mapping = {
        'Task ID': 'id',
        'Description': 'description',
        'Owner': 'owner',
        'Coverage Status': 'coverage_status',
        'Status': 'status',
        'Test Type': 'test_type',
        'Dependencies': 'dependencies',
    }
    
    # Convert test plan to list and extract data
    tasks_data = []
    for task in test_plan:
        task_dict = _extract_task_data(task)
        tasks_data.append(task_dict)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Plan"
    
    # Write headers if requested
    if include_headers:
        header_row = []
        for column in columns:
            header_row.append(column)
        ws.append(header_row)
    
    # Write task data rows
    for task_dict in tasks_data:
        row = []
        for column in columns:
            key = column_mapping.get(column, column.lower().replace(' ', '_'))
            value = task_dict.get(key, '')
            # Convert to string and handle None values
            row.append(str(value) if value is not None else '')
        ws.append(row)
    
    # Format worksheet if requested
    if format_sheet and include_headers:
        num_rows = len(tasks_data)
        num_cols = len(columns)
        _format_worksheet(ws, num_rows, num_cols)
    
    # Save workbook
    wb.save(output_path)
    
    return output_path


def export_to_xlsx_simple(
    test_plan: Iterable[Dict[str, Any]],
    output_path: Union[str, Path],
) -> Path:
    """
    Simple XLSX export function with minimal columns (Task ID, Description, Owner, Coverage Status).
    
    This is a convenience function that uses only the essential columns as specified
    in the microtask requirements and matches the example format.
    
    Args:
        test_plan: Iterable of task dictionaries with keys: id, description, owner, coverage_status
        output_path: Path where the XLSX file should be written
        
    Returns:
        Path object pointing to the created XLSX file
        
    Raises:
        ImportError: If openpyxl is not installed
        
    Example:
        >>> test_plan = [
        ...     {'id': 'task-1', 'description': 'Write unit tests', 'owner': 'john.doe',
        ...      'coverage_status': 'not_started'},
        ...     {'id': 'task-2', 'description': 'Integration tests', 'owner': 'jane.smith',
        ...      'coverage_status': 'in_progress'},
        ... ]
        >>> export_to_xlsx_simple(test_plan, 'test_plan.xlsx')
        Path('test_plan.xlsx')
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for XLSX export. Install it with: pip install openpyxl"
        )
    
    output_path = Path(output_path)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Plan"
    
    # Write headers
    ws.append(['Task ID', 'Description', 'Owner', 'Coverage Status'])
    
    # Write task data rows
    for task in test_plan:
        ws.append([
            task.get('id', ''),
            task.get('description', ''),
            task.get('owner', ''),
            task.get('coverage_status', ''),
        ])
    
    # Format worksheet
    _format_worksheet(ws, len(list(test_plan)), 4)
    
    # Save workbook
    wb.save(output_path)
    
    return output_path


def export_plan_from_db(
    dal: Any,  # TestTaskDAL type
    output_path: Union[str, Path],
    include_headers: bool = True,
    columns: Optional[List[str]] = None,
    format_sheet: bool = True,
) -> Path:
    """
    Export test plan directly from database using DAL.
    
    This function fetches tasks from the database and exports them to XLSX.
    
    Args:
        dal: TestTaskDAL instance to fetch tasks from database
        output_path: Path where the XLSX file should be written
        include_headers: Whether to include column headers (default: True)
        columns: Optional list of column names to include
        format_sheet: Whether to apply formatting (default: True)
        
    Returns:
        Path object pointing to the created XLSX file
        
    Raises:
        ImportError: If openpyxl is not installed
        
    Example:
        >>> from database.data_access_layer import TestTaskDAL
        >>> from database.postgresql_setup import get_sessionmaker
        >>> 
        >>> sessionmaker = get_sessionmaker()
        >>> session = sessionmaker()
        >>> dal = TestTaskDAL(session)
        >>> 
        >>> export_plan_from_db(dal, 'test_plan.xlsx')
        Path('test_plan.xlsx')
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for XLSX export. Install it with: pip install openpyxl"
        )
    
    # Get tasks from database
    tasks = dal.plan_tasks()  # Returns ordered list of TestTaskModel
    
    return export_to_xlsx(
        tasks,
        output_path,
        include_headers=include_headers,
        columns=columns,
        format_sheet=format_sheet,
    )


__all__ = [
    'export_to_xlsx',
    'export_to_xlsx_simple',
    'export_plan_from_db',
    'OPENPYXL_AVAILABLE',
]

