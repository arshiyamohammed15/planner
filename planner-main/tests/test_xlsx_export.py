"""
Unit tests for XLSX export functionality.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

from export.xlsx_export import OPENPYXL_AVAILABLE, export_to_xlsx, export_to_xlsx_simple
from tasks.test_task_model import CoverageStatus, TaskStatus, TestTask, TestType

# Skip all tests if openpyxl is not available
pytestmark = pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl is not installed")


def _safe_cleanup(path: Path) -> None:
    """Safely cleanup a file, ignoring permission errors on Windows."""
    try:
        if path.exists():
            path.unlink()
    except (PermissionError, OSError):
        # File might still be locked on Windows, ignore cleanup errors
        pass


def test_export_to_xlsx_simple():
    """Test simple XLSX export matching microtask example."""
    test_plan = [
        {
            'id': 'task-1',
            'description': 'Write unit tests',
            'owner': 'john.doe',
            'coverage_status': 'not_started',
        },
        {
            'id': 'task-2',
            'description': 'Integration tests',
            'owner': 'jane.smith',
            'coverage_status': 'in_progress',
        },
    ]
    
    # Use tempfile.mktemp to avoid file locking issues
    import tempfile
    output_path = Path(tempfile.mktemp(suffix='.xlsx'))
    
    try:
        export_to_xlsx_simple(test_plan, output_path)
        
        # Verify file was created
        assert output_path.exists()
        
        # Read and verify XLSX content using openpyxl
        from openpyxl import load_workbook
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Check headers
        assert ws.cell(row=1, column=1).value == 'Task ID'
        assert ws.cell(row=1, column=2).value == 'Description'
        assert ws.cell(row=1, column=3).value == 'Owner'
        assert ws.cell(row=1, column=4).value == 'Coverage Status'
        
        # Check first data row
        assert ws.cell(row=2, column=1).value == 'task-1'
        assert ws.cell(row=2, column=2).value == 'Write unit tests'
        assert ws.cell(row=2, column=3).value == 'john.doe'
        assert ws.cell(row=2, column=4).value == 'not_started'
        
        # Check second data row
        assert ws.cell(row=3, column=1).value == 'task-2'
        assert ws.cell(row=3, column=2).value == 'Integration tests'
        assert ws.cell(row=3, column=3).value == 'jane.smith'
        assert ws.cell(row=3, column=4).value == 'in_progress'
        
        wb.close()
    finally:
        # Cleanup
        _safe_cleanup(output_path)


def test_export_to_xlsx_with_testtask_objects():
    """Test XLSX export with TestTask objects."""
    test_plan = [
        TestTask(
            id='task-1',
            description='Write unit tests',
            test_type=TestType.UNIT,
            owner='john.doe',
            status=TaskStatus.PENDING,
            coverage_status=CoverageStatus.NOT_STARTED,
            dependencies=[],
        ),
        TestTask(
            id='task-2',
            description='Integration tests',
            test_type=TestType.INTEGRATION,
            owner='jane.smith',
            status=TaskStatus.IN_PROGRESS,
            coverage_status=CoverageStatus.IN_PROGRESS,
            dependencies=['task-1'],
        ),
    ]
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        output_path = Path(tmp.name)
    
    try:
        export_to_xlsx(test_plan, output_path)
        
        # Verify file was created
        assert output_path.exists()
        
        # Read and verify XLSX content
        from openpyxl import load_workbook
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Check headers
        assert ws.cell(row=1, column=1).value == 'Task ID'
        assert ws.cell(row=1, column=2).value == 'Description'
        assert ws.cell(row=1, column=3).value == 'Owner'
        assert ws.cell(row=1, column=4).value == 'Coverage Status'
        assert ws.cell(row=1, column=5).value == 'Status'
        assert ws.cell(row=1, column=6).value == 'Test Type'
        assert ws.cell(row=1, column=7).value == 'Dependencies'
        
        # Check first data row
        assert ws.cell(row=2, column=1).value == 'task-1'
        assert ws.cell(row=2, column=2).value == 'Write unit tests'
        assert ws.cell(row=2, column=3).value == 'john.doe'
        assert ws.cell(row=2, column=4).value == 'not_started'
        assert ws.cell(row=2, column=5).value == 'pending'
        assert ws.cell(row=2, column=6).value == 'unit'
        assert ws.cell(row=2, column=7).value == ''
        
        # Check second data row
        assert ws.cell(row=3, column=1).value == 'task-2'
        assert ws.cell(row=3, column=2).value == 'Integration tests'
        assert ws.cell(row=3, column=3).value == 'jane.smith'
        assert ws.cell(row=3, column=4).value == 'in_progress'
        assert ws.cell(row=3, column=5).value == 'in_progress'
        assert ws.cell(row=3, column=6).value == 'integration'
        assert ws.cell(row=3, column=7).value == 'task-1'
        
        wb.close()
    finally:
        # Cleanup - ensure workbook is closed before deleting
        try:
            if output_path.exists():
                output_path.unlink()
        except PermissionError:
            # File might still be locked, try again after a brief delay
            import time
            time.sleep(0.1)
            if output_path.exists():
                output_path.unlink()


def test_export_to_xlsx_with_dicts():
    """Test XLSX export with dictionary objects."""
    test_plan = [
        {
            'id': 'task-1',
            'description': 'Write unit tests',
            'owner': 'john.doe',
            'coverage_status': 'not_started',
            'status': 'pending',
            'test_type': 'unit',
            'dependencies': '',
        },
        {
            'id': 'task-2',
            'description': 'Integration tests',
            'owner': 'jane.smith',
            'coverage_status': 'in_progress',
            'status': 'in_progress',
            'test_type': 'integration',
            'dependencies': 'task-1',
        },
    ]
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        output_path = Path(tmp.name)
    
    try:
        export_to_xlsx(test_plan, output_path)
        
        # Verify file was created
        assert output_path.exists()
        
        # Read and verify XLSX content
        from openpyxl import load_workbook
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Check that we have headers and data
        assert ws.cell(row=1, column=1).value == 'Task ID'
        assert ws.cell(row=2, column=1).value == 'task-1'
        assert ws.cell(row=3, column=1).value == 'task-2'
        
        wb.close()
    finally:
        # Cleanup - ensure workbook is closed before deleting
        try:
            if output_path.exists():
                output_path.unlink()
        except PermissionError:
            # File might still be locked, try again after a brief delay
            import time
            time.sleep(0.1)
            if output_path.exists():
                output_path.unlink()


def test_export_to_xlsx_custom_columns():
    """Test XLSX export with custom column selection."""
    test_plan = [
        {
            'id': 'task-1',
            'description': 'Write unit tests',
            'owner': 'john.doe',
            'coverage_status': 'not_started',
            'status': 'pending',
            'test_type': 'unit',
        },
    ]
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        output_path = Path(tmp.name)
    
    try:
        custom_columns = ['Task ID', 'Description', 'Owner']
        export_to_xlsx(test_plan, output_path, columns=custom_columns)
        
        # Verify file was created
        assert output_path.exists()
        
        # Read and verify XLSX content
        from openpyxl import load_workbook
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Check headers
        assert ws.cell(row=1, column=1).value == 'Task ID'
        assert ws.cell(row=1, column=2).value == 'Description'
        assert ws.cell(row=1, column=3).value == 'Owner'
        
        # Check data row has only 3 columns
        assert ws.cell(row=2, column=1).value == 'task-1'
        assert ws.cell(row=2, column=2).value == 'Write unit tests'
        assert ws.cell(row=2, column=3).value == 'john.doe'
        
        wb.close()
    finally:
        # Cleanup - ensure workbook is closed before deleting
        try:
            if output_path.exists():
                output_path.unlink()
        except PermissionError:
            # File might still be locked, try again after a brief delay
            import time
            time.sleep(0.1)
            if output_path.exists():
                output_path.unlink()


def test_export_to_xlsx_formatting():
    """Test that XLSX export includes formatting."""
    test_plan = [
        {
            'id': 'task-1',
            'description': 'Write unit tests',
            'owner': 'john.doe',
            'coverage_status': 'not_started',
        },
    ]
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        output_path = Path(tmp.name)
    
    try:
        export_to_xlsx(test_plan, output_path, format_sheet=True)
        
        # Verify file was created
        assert output_path.exists()
        
        # Read and verify formatting
        from openpyxl import load_workbook
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Check header formatting
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.fill is not None  # Should have fill
        assert header_cell.font.bold is True  # Should be bold
        assert header_cell.font.color is not None  # Should have color
        
        # Check that frozen panes are set
        assert ws.freeze_panes == 'A2'
        
        wb.close()
    finally:
        # Cleanup - ensure workbook is closed before deleting
        try:
            if output_path.exists():
                output_path.unlink()
        except PermissionError:
            # File might still be locked, try again after a brief delay
            import time
            time.sleep(0.1)
            if output_path.exists():
                output_path.unlink()


def test_export_to_xlsx_no_headers():
    """Test XLSX export without headers."""
    test_plan = [
        {
            'id': 'task-1',
            'description': 'Write unit tests',
            'owner': 'john.doe',
            'coverage_status': 'not_started',
        },
    ]
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        output_path = Path(tmp.name)
    
    try:
        export_to_xlsx(test_plan, output_path, include_headers=False)
        
        # Verify file was created
        assert output_path.exists()
        
        # Read and verify XLSX content
        from openpyxl import load_workbook
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Should only have data rows, no headers
        assert ws.cell(row=1, column=1).value == 'task-1'
        assert ws.cell(row=1, column=1).value != 'Task ID'  # Not a header
        
        wb.close()
    finally:
        # Cleanup - ensure workbook is closed before deleting
        try:
            if output_path.exists():
                output_path.unlink()
        except PermissionError:
            # File might still be locked, try again after a brief delay
            import time
            time.sleep(0.1)
            if output_path.exists():
                output_path.unlink()


def test_export_to_xlsx_empty_plan():
    """Test XLSX export with empty test plan."""
    test_plan = []
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        output_path = Path(tmp.name)
    
    try:
        export_to_xlsx(test_plan, output_path)
        
        # Verify file was created
        assert output_path.exists()
        
        # Read and verify XLSX content
        from openpyxl import load_workbook
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Should only have headers
        assert ws.cell(row=1, column=1).value == 'Task ID'
        # No data rows
        assert ws.cell(row=2, column=1).value is None
        
        wb.close()
    finally:
        # Cleanup - ensure workbook is closed before deleting
        try:
            if output_path.exists():
                output_path.unlink()
        except PermissionError:
            # File might still be locked, try again after a brief delay
            import time
            time.sleep(0.1)
            if output_path.exists():
                output_path.unlink()


def test_export_to_xlsx_none_values():
    """Test XLSX export handles None values correctly."""
    test_plan = [
        {
            'id': 'task-1',
            'description': 'Write unit tests',
            'owner': None,  # None owner
            'coverage_status': 'not_started',
            'status': 'pending',
            'test_type': 'unit',
            'dependencies': '',
        },
    ]
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        output_path = Path(tmp.name)
    
    try:
        export_to_xlsx(test_plan, output_path)
        
        # Verify file was created
        assert output_path.exists()
        
        # Read and verify XLSX content
        from openpyxl import load_workbook
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # None should be converted to empty string
        assert ws.cell(row=2, column=3).value == ''  # Owner column
        
        wb.close()
    finally:
        # Cleanup - ensure workbook is closed before deleting
        try:
            if output_path.exists():
                output_path.unlink()
        except PermissionError:
            # File might still be locked, try again after a brief delay
            import time
            time.sleep(0.1)
            if output_path.exists():
                output_path.unlink()


def test_export_to_xlsx_simple_matches_example():
    """Test that export_to_xlsx_simple matches the microtask example format."""
    test_plan = [
        {
            'id': 'task-1',
            'description': 'Write unit tests',
            'owner': 'john.doe',
            'coverage_status': 'not_started',
        },
    ]
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        output_path = Path(tmp.name)
    
    try:
        export_to_xlsx_simple(test_plan, output_path)
        
        # Verify file was created
        assert output_path.exists()
        
        # Read and verify XLSX content matches example format
        from openpyxl import load_workbook
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Should have headers matching example
        assert ws.cell(row=1, column=1).value == 'Task ID'
        assert ws.cell(row=1, column=2).value == 'Description'
        assert ws.cell(row=1, column=3).value == 'Owner'
        assert ws.cell(row=1, column=4).value == 'Coverage Status'
        
        # Should have data
        assert ws.cell(row=2, column=1).value == 'task-1'
        
        wb.close()
    finally:
        # Cleanup - ensure workbook is closed before deleting
        try:
            if output_path.exists():
                output_path.unlink()
        except PermissionError:
            # File might still be locked, try again after a brief delay
            import time
            time.sleep(0.1)
            if output_path.exists():
                output_path.unlink()

