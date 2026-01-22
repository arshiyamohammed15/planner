"""
Unit tests for export and import functionalities.

This module tests that CSV, JSON, and XLSX export functions correctly generate
files with the right data, and that import functionality correctly reads and
maps data from external sources.
"""

from __future__ import annotations

import csv
import json
import tempfile
from pathlib import Path

import pytest

from database.models import TestTaskModel
from tasks.test_task_model import CoverageStatus, TaskStatus, TestTask, TestType

# Import export functions using importlib to handle module name issues
import importlib.util
import sys
from pathlib import Path as PathLib

# Load export modules
csv_export_path = PathLib(__file__).parent.parent / "export" / "csv_export.py"
csv_export_spec = importlib.util.spec_from_file_location("csv_export", csv_export_path)
csv_export = importlib.util.module_from_spec(csv_export_spec)
sys.modules["csv_export"] = csv_export
csv_export_spec.loader.exec_module(csv_export)  # type: ignore

json_export_path = PathLib(__file__).parent.parent / "export" / "json_export.py"
json_export_spec = importlib.util.spec_from_file_location("json_export", json_export_path)
json_export = importlib.util.module_from_spec(json_export_spec)
sys.modules["json_export"] = json_export
json_export_spec.loader.exec_module(json_export)  # type: ignore

xlsx_export_path = PathLib(__file__).parent.parent / "export" / "xlsx_export.py"
xlsx_export_spec = importlib.util.spec_from_file_location("xlsx_export", xlsx_export_path)
xlsx_export = importlib.util.module_from_spec(xlsx_export_spec)
sys.modules["xlsx_export"] = xlsx_export
xlsx_export_spec.loader.exec_module(xlsx_export)  # type: ignore

# Load import module
import_test_plans_path = PathLib(__file__).parent.parent / "import" / "import_test_plans.py"
import_test_plans_spec = importlib.util.spec_from_file_location("import_test_plans", import_test_plans_path)
import_test_plans = importlib.util.module_from_spec(import_test_plans_spec)
sys.modules["import_test_plans"] = import_test_plans
import_test_plans_spec.loader.exec_module(import_test_plans)  # type: ignore

# Import functions
export_to_csv = csv_export.export_to_csv
export_to_csv_simple = csv_export.export_to_csv_simple
export_to_json = json_export.export_to_json
export_to_json_simple = json_export.export_to_json_simple
export_to_xlsx = xlsx_export.export_to_xlsx
export_to_xlsx_simple = xlsx_export.export_to_xlsx_simple
OPENPYXL_AVAILABLE = xlsx_export.OPENPYXL_AVAILABLE

import_from_csv = import_test_plans.import_from_csv
import_from_csv_simple = import_test_plans.import_from_csv_simple
import_from_json = import_test_plans.import_from_json
ImportValidationError = import_test_plans.ImportValidationError


# ============================================================================
# Test Data Fixtures
# ============================================================================

@pytest.fixture
def sample_test_tasks():
    """Create sample TestTask objects for testing."""
    return [
        TestTask(
            id='task-1',
            description='Write unit tests for Login',
            test_type=TestType.UNIT,
            owner='john.doe',
            status=TaskStatus.PENDING,
            coverage_status=CoverageStatus.NOT_STARTED,
            dependencies=[],
        ),
        TestTask(
            id='task-2',
            description='Integration tests for Login + DB',
            test_type=TestType.INTEGRATION,
            owner='jane.smith',
            status=TaskStatus.IN_PROGRESS,
            coverage_status=CoverageStatus.IN_PROGRESS,
            dependencies=['task-1'],
        ),
        TestTask(
            id='task-3',
            description='E2E tests for Login Flow',
            test_type=TestType.E2E,
            owner='qa.team',
            status=TaskStatus.PENDING,
            coverage_status=CoverageStatus.NOT_STARTED,
            dependencies=['task-2'],
        ),
    ]


@pytest.fixture
def sample_task_dicts():
    """Create sample task dictionaries for testing."""
    return [
        {
            'id': 'task-1',
            'description': 'Write unit tests',
            'owner': 'john.doe',
            'coverage_status': 'not_started',
        },
        {
            'id': 'task-2',
            'description': 'Integration tests',
            'owner': 'jane.smith',
            'coverage_status': 'in_progress',
        },
    ]


# ============================================================================
# CSV Export Tests
# ============================================================================

class TestCSVExport:
    """Tests for CSV export functionality."""

    def test_export_to_csv_simple_generates_file(self, sample_task_dicts):
        """Test that export_to_csv_simple generates a CSV file."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            export_to_csv_simple(sample_task_dicts, output_path)
            
            assert output_path.exists()
            assert output_path.suffix == '.csv'
        finally:
            output_path.unlink()

    def test_export_to_csv_simple_has_correct_headers(self, sample_task_dicts):
        """Test that CSV file has correct headers."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            export_to_csv_simple(sample_task_dicts, output_path)
            
            with open(output_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)
                assert headers == ['Task ID', 'Description', 'Owner', 'Coverage Status']
        finally:
            output_path.unlink()

    def test_export_to_csv_simple_has_correct_data(self, sample_task_dicts):
        """Test that CSV file contains correct data."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            export_to_csv_simple(sample_task_dicts, output_path)
            
            with open(output_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader)  # Skip header
                rows = list(reader)
                
                assert len(rows) == 2
                assert rows[0] == ['task-1', 'Write unit tests', 'john.doe', 'not_started']
                assert rows[1] == ['task-2', 'Integration tests', 'jane.smith', 'in_progress']
        finally:
            output_path.unlink()

    def test_export_to_csv_with_testtask_objects(self, sample_test_tasks):
        """Test exporting TestTask objects to CSV."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            export_to_csv(sample_test_tasks, output_path)
            
            assert output_path.exists()
            
            with open(output_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                
                assert len(rows) == 3
                assert rows[0]['Task ID'] == 'task-1'
                assert rows[0]['Description'] == 'Write unit tests for Login'
                assert rows[0]['Owner'] == 'john.doe'
                assert rows[0]['Test Type'] == 'unit'
        finally:
            output_path.unlink()

    def test_export_to_csv_custom_columns(self, sample_test_tasks):
        """Test CSV export with custom columns."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            custom_columns = ['Task ID', 'Description', 'Owner']
            export_to_csv(sample_test_tasks, output_path, columns=custom_columns)
            
            with open(output_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                
                assert len(rows[0]) == 3  # Only 3 columns
                assert 'Task ID' in rows[0]
                assert 'Description' in rows[0]
                assert 'Owner' in rows[0]
        finally:
            output_path.unlink()

    def test_export_to_csv_handles_none_values(self, sample_task_dicts):
        """Test that CSV export handles None values correctly."""
        sample_task_dicts[0]['owner'] = None
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            export_to_csv_simple(sample_task_dicts, output_path)
            
            with open(output_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader)  # Skip header
                rows = list(reader)
                
                # None should be converted to empty string
                assert rows[0][2] == ''  # Owner column
        finally:
            output_path.unlink()


# ============================================================================
# JSON Export Tests
# ============================================================================

class TestJSONExport:
    """Tests for JSON export functionality."""

    def test_export_to_json_simple_generates_file(self, sample_task_dicts):
        """Test that export_to_json_simple generates a JSON file."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            export_to_json_simple(sample_task_dicts, output_path)
            
            assert output_path.exists()
            assert output_path.suffix == '.json'
        finally:
            output_path.unlink()

    def test_export_to_json_simple_has_correct_structure(self, sample_task_dicts):
        """Test that JSON file has correct structure (array format)."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            export_to_json_simple(sample_task_dicts, output_path)
            
            with open(output_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
                assert isinstance(data, list)
                assert len(data) == 2
                assert data[0]['id'] == 'task-1'
                assert data[0]['description'] == 'Write unit tests'
        finally:
            output_path.unlink()

    def test_export_to_json_with_testtask_objects(self, sample_test_tasks):
        """Test exporting TestTask objects to JSON."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            export_to_json(sample_test_tasks, output_path)
            
            assert output_path.exists()
            
            with open(output_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                
                assert 'test_plan' in data
                assert 'exported_at' in data
                assert 'total_tasks' in data
                assert data['total_tasks'] == 3
                assert len(data['test_plan']) == 3
        finally:
            output_path.unlink()

    def test_export_to_json_readable_format(self, sample_task_dicts):
        """Test that JSON is properly formatted and readable."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            export_to_json_simple(sample_task_dicts, output_path, indent=4)
            
            with open(output_path, 'r', encoding='utf-8') as f:
                content = f.read()
                
                # Should be valid JSON
                data = json.loads(content)
                assert isinstance(data, list)
                
                # Should be formatted (contain newlines)
                assert '\n' in content
        finally:
            output_path.unlink()


# ============================================================================
# XLSX Export Tests
# ============================================================================

@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl is not installed")
class TestXLSXExport:
    """Tests for XLSX export functionality."""

    def test_export_to_xlsx_simple_generates_file(self, sample_task_dicts):
        """Test that export_to_xlsx_simple generates an XLSX file."""
        import tempfile
        output_path = Path(tempfile.mktemp(suffix='.xlsx'))
        
        try:
            export_to_xlsx_simple(sample_task_dicts, output_path)
            
            assert output_path.exists()
            assert output_path.suffix == '.xlsx'
        finally:
            _safe_cleanup(output_path)

    def test_export_to_xlsx_simple_has_correct_headers(self, sample_task_dicts):
        """Test that XLSX file has correct headers."""
        import tempfile
        from openpyxl import load_workbook
        
        output_path = Path(tempfile.mktemp(suffix='.xlsx'))
        
        try:
            export_to_xlsx_simple(sample_task_dicts, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            assert ws.cell(row=1, column=1).value == 'Task ID'
            assert ws.cell(row=1, column=2).value == 'Description'
            assert ws.cell(row=1, column=3).value == 'Owner'
            assert ws.cell(row=1, column=4).value == 'Coverage Status'
            
            wb.close()
        finally:
            _safe_cleanup(output_path)

    def test_export_to_xlsx_simple_has_correct_data(self, sample_task_dicts):
        """Test that XLSX file contains correct data."""
        import tempfile
        from openpyxl import load_workbook
        
        output_path = Path(tempfile.mktemp(suffix='.xlsx'))
        
        try:
            export_to_xlsx_simple(sample_task_dicts, output_path)
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            assert ws.cell(row=2, column=1).value == 'task-1'
            assert ws.cell(row=2, column=2).value == 'Write unit tests'
            assert ws.cell(row=2, column=3).value == 'john.doe'
            assert ws.cell(row=2, column=4).value == 'not_started'
            
            wb.close()
        finally:
            _safe_cleanup(output_path)

    def test_export_to_xlsx_with_testtask_objects(self, sample_test_tasks):
        """Test exporting TestTask objects to XLSX."""
        import tempfile
        from openpyxl import load_workbook
        
        output_path = Path(tempfile.mktemp(suffix='.xlsx'))
        
        try:
            export_to_xlsx(sample_test_tasks, output_path)
            
            assert output_path.exists()
            
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Check headers
            assert ws.cell(row=1, column=1).value == 'Task ID'
            assert ws.cell(row=1, column=7).value == 'Dependencies'
            
            # Check data
            assert ws.cell(row=2, column=1).value == 'task-1'
            assert ws.cell(row=3, column=7).value == 'task-1'  # Dependencies
            
            wb.close()
        finally:
            _safe_cleanup(output_path)


def _safe_cleanup(path: Path) -> None:
    """Safely cleanup a file, ignoring permission errors on Windows."""
    try:
        if path.exists():
            path.unlink()
    except (PermissionError, OSError):
        pass


# ============================================================================
# CSV Import Tests
# ============================================================================

class TestCSVImport:
    """Tests for CSV import functionality."""

    def test_import_from_csv_simple_reads_file(self, sample_task_dicts):
        """Test that import_from_csv_simple reads a CSV file correctly."""
        # Create CSV file
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
            writer = csv.writer(tmp)
            writer.writerow(['Task ID', 'Description', 'Owner', 'Coverage Status'])
            for task in sample_task_dicts:
                writer.writerow([
                    task['id'],
                    task['description'],
                    task['owner'],
                    task['coverage_status'],
                ])
        
        try:
            tasks = import_from_csv_simple(output_path)
            
            assert len(tasks) == 2
            assert tasks[0]['id'] == 'task-1'
            assert tasks[0]['description'] == 'Write unit tests'
        finally:
            output_path.unlink()

    def test_import_from_csv_reads_testtask_objects(self):
        """Test that import_from_csv creates TestTask objects."""
        # Create CSV file
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
            writer = csv.writer(tmp)
            writer.writerow(['Task ID', 'Description', 'Owner', 'Coverage Status', 'Status', 'Test Type'])
            writer.writerow(['task-1', 'Write unit tests', 'john.doe', 'not_started', 'pending', 'unit'])
            writer.writerow(['task-2', 'Integration tests', 'jane.smith', 'in_progress', 'in_progress', 'integration'])
        
        try:
            tasks = import_from_csv(output_path)
            
            assert len(tasks) == 2
            assert isinstance(tasks[0], TestTask)
            assert tasks[0].id == 'task-1'
            assert tasks[0].test_type == TestType.UNIT
            assert tasks[1].test_type == TestType.INTEGRATION
        finally:
            output_path.unlink()

    def test_import_from_csv_validates_data(self):
        """Test that import_from_csv validates imported data."""
        # Create CSV with invalid data
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
            writer = csv.writer(tmp)
            writer.writerow(['Task ID', 'Description', 'Test Type'])
            writer.writerow(['', 'Test task', 'unit'])  # Missing ID
            writer.writerow(['task-2', 'Test task', 'invalid_type'])  # Invalid test type
        
        try:
            # Should raise validation error for missing ID
            with pytest.raises(ImportValidationError, match="Task 'id' is required"):
                import_from_csv(output_path)
        finally:
            output_path.unlink()

    def test_import_from_csv_handles_dependencies(self):
        """Test that import_from_csv correctly handles dependencies."""
        # Create CSV with dependencies
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
            writer = csv.writer(tmp)
            writer.writerow(['Task ID', 'Description', 'Dependencies'])
            writer.writerow(['task-1', 'First task', ''])
            writer.writerow(['task-2', 'Second task', 'task-1'])
        
        try:
            tasks = import_from_csv(output_path)
            
            assert len(tasks) == 2
            assert tasks[0].dependencies == []
            assert tasks[1].dependencies == ['task-1']
        finally:
            output_path.unlink()


# ============================================================================
# JSON Import Tests
# ============================================================================

class TestJSONImport:
    """Tests for JSON import functionality."""

    def test_import_from_json_reads_list_format(self):
        """Test that import_from_json reads list format JSON."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
            json.dump([
                {
                    'id': 'task-1',
                    'description': 'Write unit tests',
                    'owner': 'john.doe',
                    'coverage_status': 'not_started',
                    'test_type': 'unit',
                    'status': 'pending',
                    'dependencies': [],
                },
                {
                    'id': 'task-2',
                    'description': 'Integration tests',
                    'owner': 'jane.smith',
                    'coverage_status': 'in_progress',
                    'test_type': 'integration',
                    'status': 'in_progress',
                    'dependencies': ['task-1'],
                },
            ], tmp, indent=2)
        
        try:
            tasks = import_from_json(output_path)
            
            assert len(tasks) == 2
            assert isinstance(tasks[0], TestTask)
            assert tasks[0].id == 'task-1'
            assert tasks[0].test_type == TestType.UNIT
            assert tasks[1].dependencies == ['task-1']
        finally:
            output_path.unlink()

    def test_import_from_json_reads_test_plan_format(self):
        """Test that import_from_json reads test_plan format JSON."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
            json.dump({
                'test_plan': [
                    {
                        'id': 'task-1',
                        'description': 'Write unit tests',
                        'owner': 'john.doe',
                        'coverage_status': 'not_started',
                        'test_type': 'unit',
                    }
                ],
                'exported_at': '2026-01-15T14:00:00',
                'total_tasks': 1,
            }, tmp, indent=2)
        
        try:
            tasks = import_from_json(output_path)
            
            assert len(tasks) == 1
            assert tasks[0].id == 'task-1'
        finally:
            output_path.unlink()

    def test_import_from_json_validates_data(self):
        """Test that import_from_json validates imported data."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
            json.dump([
                {
                    'id': '',  # Invalid: empty ID
                    'description': 'Test task',
                }
            ], tmp, indent=2)
        
        try:
            with pytest.raises(ImportValidationError, match="Task 'id' is required"):
                import_from_json(output_path)
        finally:
            output_path.unlink()


# ============================================================================
# Round-Trip Tests (Export then Import)
# ============================================================================

class TestExportImportRoundTrip:
    """Tests for round-trip export/import functionality."""

    def test_csv_export_import_round_trip(self, sample_test_tasks):
        """Test that data exported to CSV can be imported back correctly."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
        
        try:
            # Export to CSV
            export_to_csv(sample_test_tasks, output_path)
            
            # Import from CSV
            imported_tasks = import_from_csv(output_path)
            
            # Verify data integrity
            assert len(imported_tasks) == len(sample_test_tasks)
            for original, imported in zip(sample_test_tasks, imported_tasks):
                assert imported.id == original.id
                assert imported.description == original.description
                assert imported.test_type == original.test_type
                assert imported.owner == original.owner
                assert imported.status == original.status
                assert imported.coverage_status == original.coverage_status
        finally:
            output_path.unlink()

    def test_json_export_import_round_trip(self, sample_test_tasks):
        """Test that data exported to JSON can be imported back correctly."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            # Export to JSON
            export_to_json(sample_test_tasks, output_path)
            
            # Import from JSON
            imported_tasks = import_from_json(output_path)
            
            # Verify data integrity
            assert len(imported_tasks) == len(sample_test_tasks)
            for original, imported in zip(sample_test_tasks, imported_tasks):
                assert imported.id == original.id
                assert imported.description == original.description
                assert imported.test_type == original.test_type
                assert imported.owner == original.owner
        finally:
            output_path.unlink()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl is not installed")
    def test_xlsx_export_csv_import_round_trip(self, sample_task_dicts):
        """Test that data exported to XLSX can be imported via CSV format."""
        import tempfile
        from openpyxl import load_workbook
        
        xlsx_path = Path(tempfile.mktemp(suffix='.xlsx'))
        csv_path = Path(tempfile.mktemp(suffix='.csv'))
        
        try:
            # Export to XLSX
            export_to_xlsx_simple(sample_task_dicts, xlsx_path)
            
            # Convert XLSX to CSV format for import
            wb = load_workbook(xlsx_path)
            ws = wb.active
            
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            
            wb.close()
            
            # Import from CSV
            imported_tasks = import_from_csv_simple(csv_path)
            
            # Verify data integrity
            assert len(imported_tasks) == len(sample_task_dicts)
            assert imported_tasks[0]['id'] == sample_task_dicts[0]['id']
            assert imported_tasks[0]['description'] == sample_task_dicts[0]['description']
        finally:
            _safe_cleanup(xlsx_path)
            _safe_cleanup(csv_path)

    def test_csv_simple_export_import_round_trip(self, sample_task_dicts):
        """Test round-trip with simple CSV format."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
        
        try:
            # Export
            export_to_csv_simple(sample_task_dicts, output_path)
            
            # Import
            imported = import_from_csv_simple(output_path)
            
            # Verify
            assert len(imported) == len(sample_task_dicts)
            assert imported[0]['id'] == sample_task_dicts[0]['id']
            assert imported[0]['description'] == sample_task_dicts[0]['description']
            assert imported[0]['owner'] == sample_task_dicts[0]['owner']
            assert imported[0]['coverage_status'] == sample_task_dicts[0]['coverage_status']
        finally:
            output_path.unlink()

    def test_json_simple_export_import_round_trip(self, sample_task_dicts):
        """Test round-trip with simple JSON format."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            # Export
            export_to_json_simple(sample_task_dicts, output_path)
            
            # Import
            imported = import_from_json(output_path)
            
            # Verify
            assert len(imported) == len(sample_task_dicts)
            assert imported[0].id == sample_task_dicts[0]['id']
            assert imported[0].description == sample_task_dicts[0]['description']
        finally:
            output_path.unlink()


# ============================================================================
# Data Mapping Tests
# ============================================================================

class TestDataMapping:
    """Tests for data mapping between export and import formats."""

    def test_export_preserves_enum_values(self, sample_test_tasks):
        """Test that enum values are correctly exported and can be imported."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
        
        try:
            # Export
            export_to_csv(sample_test_tasks, output_path)
            
            # Import
            imported = import_from_csv(output_path)
            
            # Verify enum values are preserved
            assert imported[0].test_type == TestType.UNIT
            assert imported[1].test_type == TestType.INTEGRATION
            assert imported[0].status == TaskStatus.PENDING
            assert imported[1].status == TaskStatus.IN_PROGRESS
            assert imported[0].coverage_status == CoverageStatus.NOT_STARTED
            assert imported[1].coverage_status == CoverageStatus.IN_PROGRESS
        finally:
            output_path.unlink()

    def test_export_preserves_dependencies(self, sample_test_tasks):
        """Test that dependencies are correctly exported and imported."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
        
        try:
            # Export
            export_to_json(sample_test_tasks, output_path)
            
            # Import
            imported = import_from_json(output_path)
            
            # Verify dependencies
            assert imported[0].dependencies == []
            assert imported[1].dependencies == ['task-1']
            assert imported[2].dependencies == ['task-2']
        finally:
            output_path.unlink()

    def test_export_handles_empty_optional_fields(self):
        """Test that empty optional fields are handled correctly."""
        tasks = [
            TestTask(
                id='task-1',
                description='Test task',
                test_type=TestType.UNIT,
                owner=None,  # No owner
                dependencies=[],  # No dependencies
            )
        ]
        
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
        
        try:
            # Export
            export_to_csv(tasks, output_path)
            
            # Import
            imported = import_from_csv(output_path)
            
            # Verify
            assert imported[0].owner is None or imported[0].owner == ''
            assert imported[0].dependencies == []
        finally:
            output_path.unlink()

    def test_import_maps_alternative_column_names(self):
        """Test that import correctly maps alternative column names."""
        # Create CSV with alternative column names
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
            writer = csv.writer(tmp)
            writer.writerow(['task_id', 'desc', 'assignee', 'coverage'])
            writer.writerow(['task-1', 'Test task', 'john.doe', 'not_started'])
        
        try:
            tasks = import_from_csv(output_path)
            
            assert len(tasks) == 1
            assert tasks[0].id == 'task-1'
            assert tasks[0].description == 'Test task'
            assert tasks[0].owner == 'john.doe'
        finally:
            output_path.unlink()


# ============================================================================
# Error Handling Tests
# ============================================================================

class TestExportImportErrorHandling:
    """Tests for error handling in export and import functions."""

    def test_export_to_csv_file_not_writable(self, sample_task_dicts):
        """Test error handling when CSV file cannot be written."""
        # Try to write to a directory (should fail)
        invalid_path = Path('/invalid/path/test.csv')
        
        with pytest.raises((OSError, PermissionError, FileNotFoundError)):
            export_to_csv_simple(sample_task_dicts, invalid_path)

    def test_import_from_csv_file_not_found(self):
        """Test error handling when CSV file doesn't exist."""
        with pytest.raises(FileNotFoundError):
            import_from_csv('nonexistent.csv')

    def test_import_from_json_file_not_found(self):
        """Test error handling when JSON file doesn't exist."""
        with pytest.raises(FileNotFoundError):
            import_from_json('nonexistent.json')

    def test_import_from_csv_invalid_format(self):
        """Test error handling for invalid CSV format."""
        # Create CSV without headers (empty file)
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='') as tmp:
            output_path = Path(tmp.name)
            # Write empty file or file with just data rows
            writer = csv.writer(tmp)
            writer.writerow(['task-1', 'Test task'])  # Data row without header
        
        try:
            # DictReader will create fieldnames from first row if no header
            # So this might not raise an error, but will have validation issues
            # Let's test with an actually empty file instead
            with open(output_path, 'w', newline='') as f:
                pass  # Empty file
            
            # Empty file should raise an error
            with pytest.raises((ValueError, StopIteration, KeyError)):
                import_from_csv(output_path)
        finally:
            output_path.unlink()

    def test_import_from_json_invalid_format(self):
        """Test error handling for invalid JSON format."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
            tmp.write('invalid json content')
        
        try:
            with pytest.raises(json.JSONDecodeError):
                import_from_json(output_path)
        finally:
            output_path.unlink()

    def test_import_from_json_invalid_structure(self):
        """Test error handling for invalid JSON structure."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
            output_path = Path(tmp.name)
            json.dump({'invalid': 'structure'}, tmp)
        
        try:
            with pytest.raises(ValueError, match="JSON must be a list"):
                import_from_json(output_path)
        finally:
            output_path.unlink()

